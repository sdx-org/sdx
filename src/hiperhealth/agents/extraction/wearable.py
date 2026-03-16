"""
title: Wearable data module for extracting wearable data.
"""

from __future__ import annotations

import csv
import io
import json
import tempfile

from abc import ABC, abstractmethod
from pathlib import Path
from typing import IO, Any, ClassVar, Generic, Literal, TypeVar, Union, cast

import magic

from hiperhealth.utils import is_float


class WearableDataExtractorError(Exception):
    """
    title: Base class for wearable data file errors.
    """

    ...


class FileProcessingError(WearableDataExtractorError):
    """
    title: File Processing Exception.
    """

    ...


T = TypeVar('T')
FileInput = Union[str, Path, IO[bytes], tempfile.SpooledTemporaryFile[bytes]]
FileExtension = Literal['json', 'csv', 'xlsx']
MimeType = Literal[
    'application/json',
    'text/csv',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
]


class BaseWearableDataExtractor(ABC, Generic[T]):
    """
    title: Base class for wearable data extraction.
    """

    @abstractmethod
    def extract_wearable_data(self, source: T) -> list[dict[str, object]]:
        """
        title: Implement the wearable data extraction.
        parameters:
          source:
            type: T
            description: Value for source.
        returns:
          type: list[dict[str, object]]
          description: Return value.
        """
        raise NotImplementedError(source)


class WearableDataFileExtractor(BaseWearableDataExtractor[FileInput]):
    """
    title: Wearable data file based extractor.
    attributes:
      allowed_extensions_mimetypes_map:
        type: ClassVar[dict[FileExtension, MimeType]]
        description: Value for allowed_extensions_mimetypes_map.
      _mimetype_cache:
        type: dict[str, MimeType]
        description: Value for _mimetype_cache.
      mime:
        type: magic.Magic
        description: Value for mime.
    """

    allowed_extensions_mimetypes_map: ClassVar[
        dict[FileExtension, MimeType]
    ] = {
        'json': 'application/json',
        'csv': 'text/csv',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    }

    def __init__(self) -> None:
        """
        title: Initialize caching an magic-python object.
        """
        self._mimetype_cache: dict[str, MimeType] = {}
        self.mime: magic.Magic = magic.Magic(mime=True)

    @property
    def allowed_extensions(self) -> list[FileExtension]:
        """
        title: List of supported file extensions.
        returns:
          type: list[FileExtension]
          description: Return value.
        """
        return [ext for ext in self.allowed_extensions_mimetypes_map.keys()]

    @property
    def allowed_mimetypes(self) -> list[MimeType]:
        """
        title: List of supported mimetypes.
        returns:
          type: list[MimeType]
          description: Return value.
        """
        return [
            mimetype
            for mimetype in self.allowed_extensions_mimetypes_map.values()
        ]

    def extract_wearable_data(
        self, file: FileInput
    ) -> list[dict[str, object]]:
        """
        title: Extract wearable data from file.
        parameters:
          file:
            type: FileInput
            description: Value for file.
        returns:
          type: list[dict[str, object]]
          description: Return value.
        """
        self._validate_or_raise(file)
        return self._process_file(file)

    def _process_file(self, file: FileInput) -> list[dict[str, object]]:
        if self._is_json(file):
            return self._process_json_file(file)
        elif self._is_csv(file):
            return self._process_csv_file(file)
        elif self._is_excel(file):
            return self._process_excel_file(file)
        else:
            raise FileProcessingError(
                'File could not be processed. '
                'It can be a malformed or corrupted file.'
            )

    def is_supported(self, file: FileInput) -> bool:
        """
        title: Check if file is supported.
        parameters:
          file:
            type: FileInput
            description: Value for file.
        returns:
          type: bool
          description: Return value.
        """
        if isinstance(file, (tempfile.SpooledTemporaryFile, io.BytesIO)):
            # if it's a inmemory-temp file, validate it
            return self._validate_inmemory_file(file)

        if isinstance(file, Path):
            # if it's normal file, gets its extension
            return file.suffix.replace('.', '') in self.allowed_extensions

        return self._get_mime_type(file) in self.allowed_mimetypes

    def _validate_inmemory_file(self, file: IO[bytes]) -> bool:
        try:
            file.seek(0)
            sample = file.read(10)  # read 10 bytes
            file.seek(0)
            return bool(sample)
        except Exception:
            return False

    def _validate_or_raise(self, file: FileInput) -> None:
        if not self.is_supported(file):
            raise WearableDataExtractorError(
                f'File is not valid. '
                f'Only the following extensions are accepted: '
                f'{", ".join(self.allowed_extensions)}.'
            )

    def _get_mime_type(self, file: FileInput) -> str:
        """
        title: Get MIME type of a given file input.
        parameters:
          file:
            type: FileInput
            description: Value for file.
        returns:
          type: str
          description: Return value.
        """
        cache_key = self._get_cache_key(file)

        if cache_key in self._mimetype_cache:
            # Return cached mimetype
            return self._mimetype_cache[cache_key]

        if isinstance(file, Path):
            self._mimetype_cache[cache_key] = cast(
                MimeType, self.mime.from_file(file)
            )
            return self._mimetype_cache[cache_key]
        elif isinstance(file, io.IOBase):
            head = file.read(2048)
            file.seek(0)
            self._mimetype_cache[cache_key] = cast(
                MimeType, self.mime.from_buffer(head)
            )
            return self._mimetype_cache[cache_key]
        else:
            raise TypeError(
                'Unsupported file type: must be Path or file-like object.'
            )

    def _get_cache_key(self, file: FileInput) -> str:
        cache_key: str
        if isinstance(file, Path):
            cache_key = str(file.resolve())
        else:
            cache_key = str(id(file))  # Usar o id do objeto em memória
        return cache_key

    def _is_json(self, file: FileInput) -> bool:
        if isinstance(file, (tempfile.SpooledTemporaryFile, io.BytesIO)):
            try:
                file.seek(0)
                json.loads(file.read().decode('utf-8'))
                file.seek(0)
                return True
            except json.JSONDecodeError:
                file.seek(0)
                return False
        return (
            self._get_mime_type(file)
            == self.allowed_extensions_mimetypes_map['json']
        )

    def _is_csv(self, file: FileInput) -> bool:
        if isinstance(file, (tempfile.SpooledTemporaryFile, io.BytesIO)):
            try:
                file.seek(0)
                content = file.read().decode('utf-8')
                file.seek(0)

                reader = csv.DictReader(io.StringIO(content))

                # Check if header exists and is valid (non-empty strings)
                if not reader.fieldnames or any(
                    name is None or name.strip() == ''
                    for name in reader.fieldnames
                ):
                    return False

                # Try reading at least one row and see if it's consistent
                first_row = next(reader, None)
                if first_row is None:
                    return False

                return True
            except csv.Error:
                file.seek(0)
                return False
        return (
            self._get_mime_type(file)
            == self.allowed_extensions_mimetypes_map['csv']
        )

    def _is_excel(self, file: FileInput) -> bool:
        if isinstance(file, (tempfile.SpooledTemporaryFile, io.BytesIO)):
            try:
                import openpyxl

                file.seek(0)
                wb = openpyxl.load_workbook(
                    file, read_only=True, data_only=True
                )
                wb.close()
                file.seek(0)
                return True
            except Exception:
                file.seek(0)
                return False
        return (
            self._get_mime_type(file)
            == self.allowed_extensions_mimetypes_map['xlsx']
        )

    def _process_row(self, row: dict[str, Any]) -> dict[str, object]:
        for key, value in row.items():
            if value.isnumeric():
                row[key] = int(value)
            elif is_float(value):
                row[key] = float(value)
            else:
                # remove leading and trailing whitespace
                row[key] = value.strip()
        return row

    def _process_json_file(self, file: FileInput) -> list[dict[str, object]]:
        if isinstance(file, (str, Path)):
            with open(file, 'r', encoding='utf-8') as f:
                return cast(list[dict[str, object]], json.load(f))
        else:
            file.seek(0)
            return cast(
                list[dict[str, object]],
                json.load(io.TextIOWrapper(file, encoding='utf-8')),
            )

    def _process_csv_file(self, file: FileInput) -> list[dict[str, object]]:
        if isinstance(file, (str, Path)):
            with open(file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                return [self._process_row(row) for row in reader]
        else:
            file.seek(0)
            reader = csv.DictReader(io.TextIOWrapper(file, encoding='utf-8'))
            return [self._process_row(row) for row in reader]

    def _process_excel_file(self, file: FileInput) -> list[dict[str, object]]:
        import openpyxl

        if isinstance(file, (str, Path)):
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        else:
            file.seek(0)
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

        ws = wb.active
        if not ws:
            wb.close()
            return []

        rows = ws.iter_rows(values_only=True)
        headers_tuple = next(rows, None)
        if not headers_tuple:
            wb.close()
            return []

        headers = [
            str(h).strip() if h is not None else '' for h in headers_tuple
        ]

        data = []
        for row in rows:
            row_dict = {}
            if all(cell is None for cell in row):
                continue
            for header, value in zip(headers, row):
                if header:
                    row_dict[header] = str(value) if value is not None else ''
            data.append(self._process_row(row_dict))

        wb.close()
        if not isinstance(file, (str, Path)):
            file.seek(0)
        return data
